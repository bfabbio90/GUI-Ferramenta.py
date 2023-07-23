# GUI Report Accident Tool which creates an Excel File

import openpyxl
import time
import tkinter as tk
from tkinter import ttk

janela = tk.Tk()

# Combobox: 
lista_tipos = ['Tipo 1', 'Tipo 2', 'Tipo 3', 'Tipo 4']

# Main Global Function:
def cadastro():
    global a, b, c, d, e, f, g
    a = entry_id.get()
    b = entry_nome.get()
    c = entry_cargo.get()
    d = entry_localtrab.get()
    e = entry_setor1.get()
    f = entry_setor2.get()
    g = entry_setor3.get()

# Window Title:
janela.title('Ferramenta de Cadastro de Ocorrências')

# Structure of Window:
janela.columnconfigure([0,1], weight=1)


label_id = tk.Label(text='Id:')
label_id.grid(row=1, column=0, padx=10, pady=10, sticky='nswe', columnspan=1)

entry_id = tk.Entry()
entry_id.grid(row=2, column=0, padx=10, pady=10, sticky='nswe', columnspan=1)

label_nome = tk.Label(text='Nome:')
label_nome.grid(row=1, column=1, padx=5, pady=10, sticky='nswe', columnspan=2)

entry_nome = tk.Entry()
entry_nome.grid(row=2, column=1, padx=5,pady=10, sticky='nswe', columnspan=2)

label_cargo = tk.Label(text='Cargo:')
label_cargo.grid(row=1, column=3, padx=5, pady=10, sticky='nswe', columnspan=3)

entry_cargo = tk.Entry()
entry_cargo.grid(row=2, column=3, padx=5, pady=10, sticky='nswe', columnspan=3)

label_localtrab = tk.Label(text='Local de Trabalho:')
label_localtrab.grid(row=4, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

entry_localtrab = tk.Entry()
entry_localtrab.grid(row=5, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

label_setor1 = tk.Label(text='Setor1:')
label_setor1.grid(row=7,column=0, padx=10, pady=10, sticky='nswe', columnspan=1)

entry_setor1 = tk.Entry()
entry_setor1.grid(row=8, column=0, padx=10, pady=10, sticky='nswe', columnspan=1)

label_setor2 = tk.Label(text='Setor2:')
label_setor2.grid(row=7, column=1, padx=5, pady=10, sticky='nswe', columnspan=1)

entry_setor2 = tk.Entry()
entry_setor2.grid(row=8, column=1, padx=5, pady=10, sticky='nswe', columnspan=1)

label_setor3 = tk.Label(text='Setor3:')
label_setor3.grid(row=7, column=2, padx=5, pady=10, sticky='nswe', columnspan=1)

entry_setor3 = tk.Entry()
entry_setor3.grid(row=8, column=2, padx=5, pady=10, sticky='nswe', columnspan=1)

label_setor4 = tk.Label(text='Setor4:')
label_setor4.grid(row=7, column=3, padx=5, pady=10, sticky='nswe', columnspan=1)

entry_setor4 = tk.Entry()
entry_setor4.grid(row=8, column=3, padx=5, pady=10, sticky='nswe', columnspan=1)

label_descricao = tk.Label(text='Descrição da ocorrência:')
label_descricao.grid(row=9, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

entry_descricao = tk.Entry()
entry_descricao.grid(row=10,column=0, padx=10, pady=10, sticky='nswe', columnspan=4)

label_tipo_ocorrencia = tk.Label(text='Tipo de Ocorrência:')
label_tipo_ocorrencia.grid(row=13, column=0, padx=10, pady=5, sticky='nswe', columnspan=2)

combobox_selecionar = ttk.Combobox(values=lista_tipos, width=30)
combobox_selecionar.grid(row=13, column=2, padx=20, pady=20, sticky='nswe', columnspan=2)

botao_cadastrar_ocorrencia = tk.Button(text='Cadastrar', command=cadastro)
botao_cadastrar_ocorrencia.grid(row=14, column=0, padx=10, pady=10, sticky='nswe', columnspan=4)
janela.mainloop()

# Creating DataSheet:
book = openpyxl.Workbook()

# View Page
print(book.sheetnames)

# Creating Page
book.create_sheet('Cadastro')

#Selecting Page
cadastro_page = book['Cadastro']
cadastro_page.append(['Id', 'Nome', 'Cargo' ,'Setor1','Setor2','Setor3','Setor4','Local de Trabalho']) 
cadastro_page.append([a, b, c, d, e, f, g])

# Saving Datasheet
book.save('Planilha de Cadastro.xlsx')
